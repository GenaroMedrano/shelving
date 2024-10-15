from openpyxl import Workbook, load_workbook
from datetime import date
import call_number_componant_split as call_normalize


def run_out_excel(excel_path):
    wb = load_workbook(excel_path)

    ws = wb.active
    counter = 1
    array_counter = 0
    outs_array = []

    def check_if_empty(inns):
        if inns.value:
            outs = inns.value
        else:
            outs = ""
        return outs

    miss_counter = 0
    the_last_call = ''
    # The F Column is where it's looking for the 'm' for MissShelving
    for cell in ws['F']:
        if cell.value == 'm' or cell.value == 'p' or cell.value == 's':
            i = counter - 2
            if i <= 0:
                i = 1
            while i <= counter + 2:
                if ws['E' + str(i)].value and i >= 1:
                    if i == counter:
                        outs_array.append(["*" + ws['E' + str(i)].value + "*", "*" + check_if_empty(ws['F' + str(i)]) + "*","*" + ws['I' + str(i)].value + "*"])
                        miss_counter += 1
                    else:
                        #  print(ws['D' + str(i)].value, ws['I' + str(i)].value, check_if_empty(ws['F' + str(i)]))
                        print(ws['E' + str(i)].value + "\t" + check_if_empty(ws['F' + str(i)]) + "\t" + ws['I' + str(i)].value)
                        outs_array.append([ws['E' + str(i)].value, check_if_empty(ws['F' + str(i)]), ws['I' + str(i)].value])
                the_last_call = ws['E' + str(i)].value
                i += 1
            outs_array.append(["", "", ""])
        counter += 1

    today = date.today()
    outs_array.append([str(today), str(miss_counter), "P=Pull book, S=Eye ball volume order."])
    print(miss_counter)

    wb.create_sheet('Miss Shelved')
    ws = wb['Miss Shelved']

    for i in outs_array:
        ws.append(i)

    wb.save(excel_path)

    this_day = today.strftime("%d/%m/%Y")
    log = '\n'
    last_bit = call_normalize.get_call_letters(the_last_call)

    log += this_day + "\t" + last_bit[0] + "\t" + str(miss_counter) + "\t" + str(counter - 1)

    with open(r"\\fs16.tamuk.edu\ds$\Library\circulation\Shelving Stats\process_logs.txt", "a") as myfile:
        myfile.write(log)
