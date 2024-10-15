import call_number_componant_split as call_normalize
from openpyxl import Workbook, load_workbook


def run_modify_excel(excel_path):
    file_location = excel_path
    wb = load_workbook(file_location)
    ws = wb.active

    ws.insert_rows(idx=1,amount=1)

    ws.delete_rows(1)
    ws.insert_cols(1)
    ws.insert_cols(4)
    ws.insert_cols(6)

    counter = 1

    for cell in ws['E']:
        if counter == 1:
            ws['A1'] = 'Mod Call#'
            ws['B1'] = 'Record#'
            ws['C1'] = 'Barcode'
            ws['D1'] = 'Index'
            ws['E1'] = 'CallNumber'
            ws['F1'] = 'Miss Shelved'
            ws['G1'] = 'Location'
            ws['H1'] = 'Location2'
            ws['I1'] = 'Title'
        else:
            cell_number = counter
            the_value = call_normalize.final_output(cell.value)
            ws['D' + str(cell_number)].value = cell_number
            ws['A' + str(cell_number)].value = the_value
            ws['F' + str(cell_number)].value = f'=IF(D{cell_number-1}>{cell_number},"m","")'
        counter += 1
    wb.save(file_location)
